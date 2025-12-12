#!/usr/bin/env python3
"""
Excel 파일의 실제 컬럼명을 확인하는 스크립트
"""
import sys
import os

# openpyxl을 사용하여 Excel 파일 읽기 (pandas 없이)
try:
    from openpyxl import load_workbook
    
    file_path = 'upload/2024 본예산 기초만_1.xlsx'
    
    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        sys.exit(1)
    
    print(f"=== Excel 파일 컬럼명 확인: {file_path} ===\n")
    
    # Excel 파일 열기
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # 헤더 행 찾기 (사업명, 부서명이 포함된 행)
    header_row = None
    for i in range(1, min(11, ws.max_row + 1)):
        row_values = [str(cell.value) if cell.value else '' for cell in ws[i]]
        has_project = any('사업' in str(val) for val in row_values)
        has_dept = any('부서' in str(val) or '소관' in str(val) for val in row_values)
        if has_project and has_dept:
            header_row = i
            break
    
    if header_row is None:
        header_row = 1
        print("⚠️  헤더 행을 찾지 못했습니다. 1행을 헤더로 사용합니다.\n")
    else:
        print(f"✅ 헤더 행: {header_row}행\n")
    
    # 헤더 행의 컬럼명 출력
    print("=== 컬럼명 목록 ===")
    for col_idx, cell in enumerate(ws[header_row], 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(64 + ((col_idx - 1) % 26) + 1)
        if col_idx <= 26:
            col_letter = chr(64 + col_idx)
        else:
            col_letter = chr(64 + (col_idx - 1) // 26) + chr(64 + ((col_idx - 1) % 26) + 1)
        
        value = cell.value if cell.value else ''
        print(f"{col_letter}열 ({col_idx}번째): '{value}'")
    
    print(f"\n총 {ws.max_column}개 컬럼")
    
    # 데이터 행 샘플 (헤더 다음 5개 행)
    print(f"\n=== 데이터 샘플 (헤더 다음 5개 행) ===")
    for row_idx in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        project_cell = ws.cell(row=row_idx, column=1)
        project_name = project_cell.value if project_cell.value else ''
        print(f"\n{row_idx}행: '{project_name}'")
        
        # 각 컬럼의 값 출력 (비어있지 않은 것만)
        for col_idx in range(1, min(ws.max_column + 1, 25)):  # 처음 24개 컬럼만
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                col_letter = chr(64 + col_idx) if col_idx <= 26 else ''
                header_cell = ws.cell(row=header_row, column=col_idx)
                header_name = header_cell.value if header_cell.value else ''
                print(f"  {col_letter}열 ({header_name}): {cell.value}")
    
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다.")
    print("설치 방법: pip install openpyxl")
    sys.exit(1)
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

